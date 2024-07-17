import json
import openpyxl
import sys
import time

from collections import defaultdict
from numpy import Infinity
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from qiskit import qasm2
from qiskit.circuit.library.standard_gates import SwapGate
from qiskit.converters import circuit_to_dag
from qiskit.transpiler import CouplingMap, PassManager
from qiskit.transpiler.passes import (
    Collect2qBlocks,
    ConsolidateBlocks,
    Unroll3qOrMore,
    UnitarySynthesis,
    FullAncillaAllocation,
    EnlargeWithAncilla,
    ApplyLayout,
    Decompose
)
from statistics import mean
from typing import Tuple
from utils import BACKENDS

'''LAYOUT'''
from qiskit.transpiler.passes import SabreLayout
from layout.layout_010 import Layout_010
from layout.layout_018 import Layout_018
from layout.layout_0181 import Layout_0181
from layout.layout_023 import Layout_023
from layout.layout_037 import Layout_037
from layout.layout_040 import Layout_040
from layout.layout_044 import Layout_044
from layout.layout_087 import Layout_087
from layout.layout_0871 import Layout_0871
from layout.layout_0872 import Layout_0872
from layout.layout_test import Layout_Test

'''ROUTING'''
from qiskit.transpiler.passes import SabreSwap
from routing.routing_010 import Routing_010
from routing.routing_018 import Routing_018
from routing.routing_023 import Routing_023
from routing.routing_037 import Routing_037
from routing.routing_040 import Routing_040
from routing.routing_044 import Routing_044
from routing.routing_087 import Routing_087
from routing.routing_test import Routing_Test

def initialize_circuit(filepath):
    circ = qasm2.load(filepath)
    dag = circuit_to_dag(circ)
    two_gates = dag.two_qubit_ops()
    mul_gates = dag.multi_qubit_ops()
    non_single_gates = set()

    for gate in two_gates + mul_gates:
        if gate.name == "cx":
            continue
        non_single_gates.add(gate)

    non_single_gates = list(non_single_gates)
    init_pm = PassManager([
        Unroll3qOrMore(),
        Decompose(non_single_gates)
    ])
    init_start = time.time()
    init_circ = init_pm.run(circ)
    init_end = time.time()

    return init_circ, init_start, init_end

def run(
        bench_path: str,
        backend: str,
        test_methods: list[str],
        reps: int = 5,
        objective: str = "size",
        verbose: bool = False
) -> Tuple[int, defaultdict[int], int, defaultdict[int], defaultdict[float]]:

    size_out = defaultdict(lambda: Infinity)
    depth_out = defaultdict(lambda: Infinity)
    runtime = defaultdict(float)
    using_layout = [0,0,0,0,0]
    coupling_map = CouplingMap(couplinglist=BACKENDS[backend].edges())
    coupling_map.make_symmetric()
    init_circ, init_start, init_end = initialize_circuit(bench_path)
    print(f"INPUT CIRCUIT QREGS: {init_circ.qregs}")
    print(f"INPUT CIRCUIT SIZE: {init_circ.size()}")
    print(f"INPUT CIRCUIT DEPTH: {init_circ.depth()}")
    size_in = init_circ.size()
    depth_in = init_circ.depth()

    for method in test_methods:
        print("=" * (len(method) + 4))
        print(f"| {method} |")
        print("=" * (len(method) + 4))


        print(coupling_map)
        if method.lower() == "sabre":
            trial_layout = SabreLayout(coupling_map, skip_routing=True)
            trial_routing = SabreSwap(coupling_map, heuristic="basic")
        elif method.lower() == "test":
            trial_layout = Layout_Test(coupling_map)
            trial_routing = Routing_Test(coupling_map)
        elif method.lower() == "018":
            trial_layout = Layout_018(coupling_map)
            trial_routing = Routing_018(coupling_map)
        elif method.lower() == "0181":
            trial_layout = Layout_0181(coupling_map)
            trial_routing = SabreSwap(coupling_map, heuristic="basic")
        elif method.lower() == "087":
            trial_layout = Layout_087(coupling_map)
            trial_routing = Routing_087(coupling_map)
        elif method.lower() == "0871":
            trial_layout = Layout_0871(coupling_map)
            trial_routing = SabreSwap(coupling_map, heuristic="basic")
        elif method.lower() == "0872":
            trial_layout = Layout_0872(coupling_map)
            trial_routing = SabreSwap(coupling_map, heuristic="basic")
        elif method.lower() == "037":
            trial_layout = Layout_037(coupling_map)
            trial_routing = Routing_037(coupling_map)
        elif method.lower() == "044":
            trial_layout = Layout_044(coupling_map)
            trial_routing = Routing_044(coupling_map)
        elif method.lower() == "010":
            trial_layout = Layout_010(coupling_map)
            trial_routing = Routing_010(coupling_map)
        elif method.lower() == "040":
            trial_layout = Layout_040(coupling_map)
            trial_routing = Routing_040(coupling_map)
        elif method.lower() == "023":
            trial_layout = Layout_023(coupling_map)
            trial_routing = SabreSwap(coupling_map, heuristic="basic")
        else:
            raise ValueError("No such method")

        for i in range(1, reps+1):
            print(f"method: {method}, reps = {i}")
            if method.lower() == "sabre":
                pm = PassManager([
                    trial_layout,
                    FullAncillaAllocation(coupling_map),
                    EnlargeWithAncilla(),
                    ApplyLayout(),
                    trial_routing,
                    Decompose(SwapGate)
                ])
            else:
                pm = PassManager([
                    trial_layout,
                    FullAncillaAllocation(coupling_map),
                    EnlargeWithAncilla(),
                    ApplyLayout(),
                    trial_routing,
                    Decompose(SwapGate)
                ])
            start = time.time()
            result_circ = pm.run(init_circ)
            # if method.lower() != "sabre" : 
            #     using_layout[trial_layout.using_mapping_number] += 1
            #     print("In main : ",trial_layout.using_mapping_number)
            end = time.time()
            size = result_circ.size()
            depth = result_circ.depth()
            
            t = (init_end - init_start) + (end - start)
            if objective.lower() == "size" and size >= size_out[method]:
                continue
            if objective.lower() == "depth" and depth >= depth_out[method]:
                continue
            size_out[method] = size
            depth_out[method] = depth
            runtime[method] = t

            if verbose:
                print(f"Updating best record of {method} ...")
                print(f"    size: {size}")
                print(f"    depth: {depth}")
                print(f"    runtime: {t}")

    return size_in, size_out, depth_in, depth_out, runtime

def export_excel(
        excel_path: str,
        test_methods: list[str],
        size_ins: defaultdict[int],
        size_outs: defaultdict[dict],
        depth_ins: defaultdict[int],
        depth_outs: defaultdict[dict],
        runtimes: defaultdict[dict]
) -> None:

    filepath = (Path(__file__).parent).joinpath(excel_path)

    # If the file is already existed, delete it first, then create new one
    if filepath.is_file():
        filepath.unlink()

    wb = openpyxl.Workbook()
    wb.remove(wb.worksheets[0])
    wb.create_sheet("Size")
    wb.create_sheet("Depth")

    for sheet in wb.worksheets:
        sheet['A1'] = "Benchmark"
        sheet['B1'] = sheet.title + "_in"

        for idx, method in enumerate(test_methods, start=1):
            col_idx = idx * 3
            prefix = method.upper()
            sheet.cell(row=1, column=col_idx).value = f"{prefix}_{sheet.title}_out"
            sheet.cell(row=1, column=col_idx+1).value = f"{prefix}_{sheet.title}_ratio"
            sheet.cell(row=1, column=col_idx+2).value = f"{prefix}_runtime"
        
        for row_idx, bench_name in enumerate(size_ins.keys(), start=2):
            sheet.cell(row=row_idx, column=1).value = bench_name

            if sheet.title.lower() == "size":
                sheet.cell(row=row_idx, column=2).value = size_ins[bench_name]

                for idx, method in enumerate(test_methods, start=1):
                    col_idx = idx * 3
                    ratio = round(size_outs[bench_name][method]/size_ins[bench_name], 3)
                    runtime = round(runtimes[bench_name][method], 2)
                    sheet.cell(row=row_idx, column=col_idx).value = size_outs[bench_name][method]
                    sheet.cell(row=row_idx, column=col_idx+1).value = ratio
                    sheet.cell(row=row_idx, column=col_idx+2).value = runtime

            elif sheet.title.lower() == "depth":
                sheet.cell(row=row_idx, column=2).value = depth_ins[bench_name]

                for idx, method in enumerate(test_methods, start=1):
                    col_idx = idx * 3
                    ratio = round(depth_outs[bench_name][method]/depth_ins[bench_name], 3)
                    runtime = round(runtimes[bench_name][method], 2)
                    sheet.cell(row=row_idx, column=col_idx).value = depth_outs[bench_name][method]
                    sheet.cell(row=row_idx, column=col_idx+1).value = ratio
                    sheet.cell(row=row_idx, column=col_idx+2).value = runtime
    
    # Style configurations
    for sheet in wb.worksheets:
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 8
        best_records = dict.fromkeys(range(2, sheet.max_row+1), (Infinity, list()))

        for col_idx in range(3, sheet.max_column+1):
            letter = get_column_letter(col_idx)
            sheet.column_dimensions[letter].width = len(sheet[letter+"1"].value)
            
            if col_idx % 3 == 1:
                # Iterate every benchmark
                for row_idx in range(2, sheet.max_row+1):
                    curr_cell = sheet.cell(row=row_idx, column=col_idx)
                    curr_cell.font = Font(name="BIZ UDPGothic", size=14, color="E74C3C")

                    if curr_cell.value < best_records[row_idx][0]:
                        best_records[row_idx] = (curr_cell.value, [col_idx])

                    elif curr_cell.value == best_records[row_idx][0]:
                        tmp = best_records[row_idx][1]
                        tmp.append(col_idx)
                        best_records[row_idx] = (curr_cell.value, tmp)
        
        # Highlight the best methods for each benchmark
        for row_idx, (_, cols) in best_records.items():
            for col_idx in cols:
                bg = PatternFill(fgColor="FDBD01", fill_type="solid")
                sheet.cell(row=row_idx, column=col_idx).fill = bg

    wb.save(filepath)
    wb.close()

def main(
        bench_folder: str,
        bench_filter: str,
        qasm_path: str,
        excel_path: str,
        backend: str,
        reps: int,
        objective: str,
        test_methods: list[str],
        output_qasm: bool,
        output_excel: bool,
        verbose: bool
) -> None:

    if reps < 0:
        raise ValueError("`reps` must be a non-zero positive integer")

    size_ins = defaultdict(int)
    depth_ins = defaultdict(int)
    size_outs = defaultdict(dict)
    depth_outs = defaultdict(dict)
    runtimes = defaultdict(dict)

    print("+++++++++++++++++++++++++++++++++++++")
    print(f"BACKEND\t\t{backend}")
    print(f"BACKEND INFO\t{BACKENDS[backend]}")
    print(f"TEST METHODS\t\t{', '.join(test_methods)}")
    total_using_layout = [0,0,0,0,0]
    for bench_path in Path(bench_folder).iterdir():
        filename = str(bench_path).split('\\')[-1].replace(".qasm", "")
        print(f"BENCHMARK\t{bench_path}")
        print("+++++++++++++++++++++++++++++++++++++")

        size_in, size_out, depth_in, depth_out, runtime = run(
            bench_path=bench_path,
            backend=backend,
            test_methods=test_methods,
            reps=reps,
            objective=objective,
            verbose=verbose
        ) 
        # print("Benchmark",bench_path,"Using Layout",layout_option)
        # total_using_layout  = [total_using_layout[i] + layout_option[i] for i in range(len(layout_option))]
        size_ins[filename] = size_in
        depth_ins[filename] = depth_in
        for method in test_methods:
            size_outs[filename][method] = size_out[method]
            depth_outs[filename][method] = depth_out[method]
            runtimes[filename][method] = runtime[method]
            print(method)
            print(f"\tsize out/in ratio: {round(size_outs[filename][method]/size_in, 3)}")
            print(f"\tdepth out/in ratio: {round(depth_outs[filename][method]/depth_in, 3)}")
            print(f"\truntime (s): {round(runtimes[filename][method], 2)}")
        print("-------------------------------------")

    print("================ FINAL RESULT ================")
    
    total_size_in = sum(size_ins.values())
    total_depth_in = sum(depth_ins.values())
    for method in test_methods:
        print(method)
        total_size_out = sum([size_outs[filename][method] for filename in size_outs.keys()])
        total_depth_out = sum([depth_outs[filename][method] for filename in depth_outs.keys()])
        avg_runtime = mean([runtimes[filename][method] for filename in runtimes.keys()])
        print(f"\taverage size out/in ratio: {round(total_size_out/total_size_in, 3)}")
        print(f"\taverage depth out/in ratio: {round(total_depth_out/total_depth_in, 3)}")
        print(f"\taverage runtime (s): {round(avg_runtime, 2)}")
    print("Total Using Layout : ",total_using_layout)
    

    finish_time = time.strftime("%c", time.localtime())
    print(f"\nFINISH TIME: {finish_time}")
    
    if output_excel:
        export_excel(
            excel_path=excel_path,
            test_methods=test_methods,
            size_ins=size_ins,
            size_outs=size_outs,
            depth_ins=depth_ins,
            depth_outs=depth_outs,
            runtimes=runtimes
        )

if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print("Expected: python main.py <test_config.json> [--verbose or -v]")
        sys.exit(1)
    
    verbose = False
    if len(sys.argv) == 3:
        if sys.argv[-1] in ("--verbose", "-v"):
            verbose = True
        else:
            print("Only support flags like '--verbose' or '-v'")
            sys.exit(1)
    
    file = open(sys.argv[1])
    config = json.load(file)
    file.close()

    main(
            bench_folder=config["bench_folder"],
            bench_filter=config["bench_filter"],
            qasm_path=config["qasm_path"],
            excel_path=config["excel_path"],
            backend=config["backend"],
            reps=config["reps"],
            objective=config["objective"],
            test_methods=config["test_methods"],
            output_qasm=config["output_qasm"],
            output_excel=config["output_excel"],
            verbose=verbose
    )